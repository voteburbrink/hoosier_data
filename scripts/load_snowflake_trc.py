"""
load_snowflake_trc.py
Convert the DLGF TRC workbooks to CSV, PUT them to the Snowflake internal stage,
and COPY INTO RAW.DLGF_TAX_DISTRICT_UNITS (KAN-153).

Run scripts/download_dlgf_trc.py first to fetch the xlsx files.

Credentials are read from environment variables (set before running):
    $env:SNOWFLAKE_USER     = "your_username"
    $env:SNOWFLAKE_PASSWORD = "your_password"
Or pass --user/--password/--passcode on the command line. MFA passcode is the
6-digit TOTP from your authenticator app (same flow as load_snowflake_parcel.py).

Usage:
    # Convert + load default years (2024, 2025):
    python load_snowflake_trc.py

    # Specific files:
    python load_snowflake_trc.py --files propertydata/trc/2025_trc_unit.xlsx

Dependencies: pip install snowflake-connector-python openpyxl

NOTE: RAW.DLGF_TAX_DISTRICT_UNITS uses CREATE TABLE IF NOT EXISTS and this script
COPYs with PURGE=FALSE. Re-running re-loads the same rows (duplicates). To refresh
a year, TRUNCATE or DELETE that budget_year first — never CREATE OR REPLACE the
RAW table (June 5 2026 incident: CREATE OR REPLACE wiped ~22M rows).
"""

import argparse
import csv
import os
import sys
import time
import tkinter as tk
from tkinter import simpledialog

import openpyxl
import snowflake.connector

ACCOUNT   = "docbdlg-sc53221"
DATABASE  = "HOOSIER_DATA"
SCHEMA    = "RAW"
WAREHOUSE = "COMPUTE_WH"
ROLE      = "ACCOUNTADMIN"
STAGE     = "@HOOSIER_DATA.RAW.GATEWAY_STAGE"
TABLE     = "HOOSIER_DATA.RAW.DLGF_TAX_DISTRICT_UNITS"

DEFAULT_FILES = [
    os.path.join("propertydata", "trc", "2024_trc_unit.xlsx"),
    os.path.join("propertydata", "trc", "2025_trc_unit.xlsx"),
]

# Source column order in the DLGF workbook == RAW table column order.
COLUMNS = [
    "YR_NBR", "CNTY_CD", "UNIT_TYPE_CD", "UNIT_CD", "UNIT_NAME",
    "FUND_CD", "FUND_LONG_NAME", "TAX_DIST_CD", "TAX_DIST_NAME",
    "CERTD_TAX_RATE_PCNT",
]


def prompt_credentials():
    root = tk.Tk()
    root.withdraw()
    user = simpledialog.askstring("Snowflake Login", "Username:", parent=root)
    if not user:
        sys.exit("Cancelled.")
    password = simpledialog.askstring("Snowflake Login", "Password:", show="*", parent=root)
    if not password:
        sys.exit("Cancelled.")
    passcode = simpledialog.askstring("Snowflake Login", "MFA passcode (6-digit):", parent=root)
    if not passcode:
        sys.exit("Cancelled.")
    root.destroy()
    return user.strip(), password, passcode.strip()


def xlsx_to_csv(xlsx_path):
    """Flatten the workbook's single sheet to CSV (header stripped). Returns CSV path."""
    csv_path = os.path.splitext(xlsx_path)[0] + ".csv"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # drop header; SKIP_HEADER also set on COPY for safety
            # Trim trailing whitespace DLGF pads fund names with
            w.writerow(["" if c is None else str(c).strip() for c in row])
            rows += 1
    print(f"  {os.path.basename(xlsx_path)} -> {os.path.basename(csv_path)} ({rows:,} rows)")
    return csv_path


def connect(user, password, passcode):
    print(f"Connecting to Snowflake ({ACCOUNT})...")
    conn = snowflake.connector.connect(
        account=ACCOUNT, user=user, password=password,
        authenticator="username_password_mfa", passcode=passcode,
        database=DATABASE, schema=SCHEMA, warehouse=WAREHOUSE, role=ROLE,
    )
    print("  Connected.\n")
    return conn


def create_table(cursor):
    print(f"Creating {TABLE} if not exists...")
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE} (
            budget_year         VARCHAR,
            county_number       VARCHAR,
            unit_type_cd        VARCHAR,
            unit_code           VARCHAR,
            unit_name           VARCHAR,
            fund_code           VARCHAR,
            fund_name           VARCHAR,
            tax_district_code   VARCHAR,
            tax_district_name   VARCHAR,
            certd_tax_rate_pct  VARCHAR
        )
    """)
    print("  Done.\n")


def put_file(cursor, local_path):
    uri = "file://" + os.path.abspath(local_path).replace("\\", "/")
    print(f"PUT: {os.path.basename(local_path)} -> {STAGE}")
    cursor.execute(f"PUT '{uri}' {STAGE} AUTO_COMPRESS=TRUE OVERWRITE=TRUE PARALLEL=4")
    for r in cursor.fetchall():
        status = r[6] if len(r) > 6 else "?"
        print(f"  {r[0]} status={status}")


def copy_into(cursor):
    print(f"\nCOPY INTO {TABLE}...")
    cursor.execute(f"""
        COPY INTO {TABLE}
        FROM {STAGE}
        PATTERN = '.*_trc_unit\\.csv.*'
        FILE_FORMAT = (
            TYPE = 'CSV'
            SKIP_HEADER = 0
            FIELD_OPTIONALLY_ENCLOSED_BY = '"'
            NULL_IF = ('')
            ERROR_ON_COLUMN_COUNT_MISMATCH = FALSE
        )
        ON_ERROR = 'CONTINUE'
        PURGE = FALSE
    """)
    rows = cursor.fetchall()
    loaded = sum(r[3] for r in rows if len(r) > 3 and r[3] is not None)
    for r in rows:
        print(f"  {os.path.basename(str(r[0]))} status={r[1]} loaded={r[3]} errors={r[5]}")
    print(f"  Rows loaded: {loaded:,}")
    return loaded


def main():
    parser = argparse.ArgumentParser(description="Load DLGF TRC workbooks into Snowflake")
    parser.add_argument("--user", default=os.environ.get("SNOWFLAKE_USER", ""))
    parser.add_argument("--password", default=os.environ.get("SNOWFLAKE_PASSWORD", ""))
    parser.add_argument("--passcode", default="", help="6-digit MFA TOTP code")
    parser.add_argument("--files", nargs="+", default=DEFAULT_FILES,
                        help="TRC xlsx file paths (default: 2024 + 2025)")
    args = parser.parse_args()

    missing = [f for f in args.files if not os.path.exists(f)]
    if missing:
        print("ERROR: file(s) not found — run download_dlgf_trc.py first:")
        for f in missing:
            print(f"  {f}")
        sys.exit(1)

    if not args.user or not args.password or not args.passcode:
        args.user, args.password, args.passcode = prompt_credentials()

    print("Converting workbooks to CSV...")
    csv_files = [xlsx_to_csv(f) for f in args.files]

    conn = connect(args.user, args.password, args.passcode)
    cursor = conn.cursor()
    try:
        create_table(cursor)
        for path in csv_files:
            put_file(cursor, path)
        copy_into(cursor)
        cursor.execute(f"SELECT budget_year, COUNT(*) FROM {TABLE} GROUP BY 1 ORDER BY 1")
        print("\nRow counts by budget_year:")
        for yr, n in cursor.fetchall():
            print(f"  {yr}: {n:,}")
    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()
